# -*- coding: utf-8 -*-
"""
main.py
交换机配置备份工具 - 主程序 (Tkinter GUI)

功能:
  1. SSH/Telnet 登录 华为 / 华三 / 锐捷 交换机备份配置
  2. 按实际主机名分区保存配置文件
  3. 批量导入 Excel/CSV/TXT 设备清单
  4. 多线程并行备份
  5. 配置对比 / 全文搜索
  6. 定时自动备份
  7. 邮件通知备份结果
  8. 设备列表持久化 / 报告导出
"""
import os
import sys
import threading
import webbrowser
import tkinter as tk
from datetime import datetime
from tkinter import ttk, filedialog, messagebox, scrolledtext

from device import Device, save_devices_json, load_devices_json, VENDOR_TYPES, VENDOR_LABELS
from backup_engine import BackupEngine
from importer import import_devices, export_template
from app_config import AppConfig
from scheduler import BackupScheduler
from notifier import send_report, build_report_html
from gui_dialogs import (DeviceDialog, SettingsDialog, CompareDialog,
                         SearchDialog, BatchEntryDialog, ImportDialog)

APP_TITLE = "交换机配置备份工具 v1.0"
APP_DIR = os.path.dirname(os.path.abspath(__file__))


# ---------- 工具函数 ----------
def ip_key(ip_str: str) -> tuple:
    """把 IP 字符串转成整数元组, 便于按段比较排序

    '10.0.0.9'  -> (10, 0, 0, 9)
    '10.0.0.10' -> (10, 0, 0, 10)   (排序后 9 < 10, 字符串比较会是 10 < 9)
    非法 IP 回退返回 (255, 255, 255, 255) 排最后
    """
    try:
        parts = ip_str.strip().split(".")
        if len(parts) != 4:
            raise ValueError
        return tuple(int(p) for p in parts)
    except (ValueError, TypeError, AttributeError):
        return (255, 255, 255, 255)


def sort_devices_by_status_and_ip(devices: list) -> list:
    """按状态和 IP 排序: 成功(IP升序) -> 失败(IP升序) -> 其余(IP升序)

    成功和失败的设备分别按 IP 从小到大排, 未备份/备份中的排在后面保持 IP 升序。
    非破坏性: 返回新列表, 原列表不变。
    """
    ok = [d for d in devices if d.status == "成功"]
    fail = [d for d in devices if d.status == "失败"]
    rest = [d for d in devices if d.status not in ("成功", "失败")]
    ok.sort(key=lambda d: ip_key(d.host))
    fail.sort(key=lambda d: ip_key(d.host))
    rest.sort(key=lambda d: ip_key(d.host))
    return ok + fail + rest


class MainWindow:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title(APP_TITLE)
        self.root.geometry("1100x720")
        self.root.minsize(900, 560)

        self.config = AppConfig()
        self.devices: list[Device] = []
        self.engine: BackupEngine = None
        self.scheduler: BackupScheduler = None
        self._backup_thread = None
        self._is_backup_running = False

        self._build_menu()
        self._build_toolbar()
        self._build_main()
        self._build_statusbar()

        self._load_devices()
        self._refresh_table()
        self._apply_scheduler()

        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

    # ====================== UI 构建 ======================
    def _build_menu(self):
        menubar = tk.Menu(self.root)

        m_file = tk.Menu(menubar, tearoff=0)
        m_file.add_command(label="新建设备清单", command=self._new_list)
        m_file.add_command(label="打开设备清单", command=self._open_list)
        m_file.add_command(label="保存设备清单", command=self._save_list)
        m_file.add_command(label="清空设备列表", command=self._clear_all)
        m_file.add_separator()
        m_file.add_command(label="批量录入(粘贴IP/账号/密码)", command=self._batch_entry)
        m_file.add_command(label="批量导入(Excel/CSV/TXT)", command=self._import_file)
        m_file.add_command(label="统一密码导入(粘贴IP+统一账号)", command=lambda: self._unified_import())
        m_file.add_command(label="导出导入模板", command=self._export_template)
        m_file.add_separator()
        m_file.add_command(label="退出", command=self._on_close)
        menubar.add_cascade(label="文件", menu=m_file)

        m_backup = tk.Menu(menubar, tearoff=0)
        m_backup.add_command(label="开始备份(全部)", command=self._backup_all)
        m_backup.add_command(label="备份选中设备", command=self._backup_selected)
        m_backup.add_command(label="停止备份", command=self._stop_backup)
        menubar.add_cascade(label="备份", menu=m_backup)

        m_tool = tk.Menu(menubar, tearoff=0)
        m_tool.add_command(label="配置对比", command=self._open_compare)
        m_tool.add_command(label="配置搜索", command=self._open_search)
        m_tool.add_command(label="打开备份目录", command=self._open_backup_dir)
        m_tool.add_command(label="导出备份报告(Excel)", command=self._export_report)
        menubar.add_cascade(label="工具", menu=m_tool)

        m_set = tk.Menu(menubar, tearoff=0)
        m_set.add_command(label="设置", command=self._open_settings)
        menubar.add_cascade(label="设置", menu=m_set)

        m_help = tk.Menu(menubar, tearoff=0)
        m_help.add_command(label="关于", command=self._about)
        menubar.add_cascade(label="帮助", menu=m_help)

        self.root.config(menu=menubar)

    def _build_toolbar(self):
        # 顶部标题栏 (蓝色横幅)
        banner = tk.Frame(self.root, bg="#1e3a8a", height=56)
        banner.pack(fill="x", side="top")
        banner.pack_propagate(False)
        tk.Label(banner, text="  交换机配置备份工具",
                 font=("Microsoft YaHei", 15, "bold"),
                 fg="white", bg="#1e3a8a").pack(side="left", padx=12)
        tk.Label(banner, text="华为 · 华三 · 锐捷  |  SSH/Telnet · 多线程 · 自动识别",
                 font=("Microsoft YaHei", 9), fg="#bfdbfe",
                 bg="#1e3a8a").pack(side="left", padx=8, pady=2)
        # 右上角统计
        self.stat_var = tk.StringVar(value="共 0 台")
        tk.Label(banner, textvariable=self.stat_var,
                 font=("Microsoft YaHei", 10, "bold"), fg="#fde047",
                 bg="#1e3a8a").pack(side="right", padx=14)

        # 工具栏 (浅灰背景)
        bar = tk.Frame(self.root, bg="#e2e8f0", height=46)
        bar.pack(fill="x", side="top")
        bar.pack_propagate(False)

        def mkbtn(parent, text, cmd, bg="#f1f5f9", fg="#1e293b"):
            return tk.Button(parent, text=text, command=cmd,
                             bg=bg, fg=fg, font=("Microsoft YaHei", 9),
                             relief="flat", padx=10, pady=4, cursor="hand2",
                             activebackground="#cbd5e1", activeforeground="#0f172a",
                             bd=0)

        # 设备管理组
        mkbtn(bar, "＋ 添加设备", self._add_device).pack(side="left", padx=(8, 2), pady=7)
        mkbtn(bar, "✎ 编辑", self._edit_device).pack(side="left", padx=2, pady=7)
        mkbtn(bar, "✖ 删除", self._del_device).pack(side="left", padx=2, pady=7)
        mkbtn(bar, "🗑 清空", self._clear_all).pack(side="left", padx=2, pady=7)
        self._vsep(bar)
        # 导入组
        mkbtn(bar, "📋 批量录入", self._batch_entry).pack(side="left", padx=2, pady=7)
        mkbtn(bar, "📂 批量导入", self._import_file).pack(side="left", padx=2, pady=7)
        mkbtn(bar, "📈 导出报告", self._export_report).pack(side="left", padx=2, pady=7)
        self._vsep(bar)
        # 备份组
        mkbtn(bar, "▶ 开始备份", self._backup_all, bg="#16a34a", fg="white").pack(side="left", padx=2, pady=7)
        mkbtn(bar, "■ 停止", self._stop_backup, bg="#dc2626", fg="white").pack(side="left", padx=2, pady=7)
        self._vsep(bar)
        # 工具组
        mkbtn(bar, "🔍 配置对比", self._open_compare).pack(side="left", padx=2, pady=7)
        mkbtn(bar, "🔎 搜索", self._open_search).pack(side="left", padx=2, pady=7)
        mkbtn(bar, "📁 备份目录", self._open_backup_dir).pack(side="left", padx=2, pady=7)
        mkbtn(bar, "⚙ 设置", self._open_settings).pack(side="right", padx=8, pady=7)

    def _vsep(self, parent):
        tk.Frame(parent, bg="#cbd5e1", width=1).pack(side="left", fill="y", padx=6, pady=8)

    def _build_main(self):
        container = tk.Frame(self.root, bg="#f8fafc")
        container.pack(fill="both", expand=True, padx=8, pady=4)

        # 表格区卡片
        table_card = tk.Frame(container, bg="white", bd=0, highlightbackground="#e2e8f0",
                              highlightthickness=1)
        table_card.pack(fill="both", expand=True)

        # 表格标题行
        th = tk.Frame(table_card, bg="#f1f5f9", height=30)
        th.pack(fill="x")
        th.pack_propagate(False)
        tk.Label(th, text="  设备列表", font=("Microsoft YaHei", 10, "bold"),
                 bg="#f1f5f9", fg="#334155").pack(side="left", padx=4)

        cols = ("name", "host", "vendor", "protocol", "port",
                "group", "status", "last_time", "message")
        headers = ["设备名称", "IP 地址", "厂商", "协议", "端口",
                   "分组", "状态", "最后备份时间", "详情"]
        widths = [130, 130, 110, 70, 60, 80, 80, 150, 260]
        tree_frame = tk.Frame(table_card, bg="white")
        tree_frame.pack(fill="both", expand=True)

        self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings")
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(side="left", fill="both", expand=True)

        sb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        sb.pack(side="right", fill="y")
        self.tree.config(yscrollcommand=sb.set)

        # 状态行配色
        self.tree.tag_configure("ok", foreground="#16a34a", background="#f0fdf4")
        self.tree.tag_configure("fail", foreground="#dc2626", background="#fef2f2")
        self.tree.tag_configure("running", foreground="#2563eb", background="#eff6ff")
        self.tree.tag_configure("pending", foreground="#64748b")
        # 隔行底色
        self.tree.tag_configure("even", background="#f8fafc")

        # 进度条
        pf = tk.Frame(self.root, bg="#f8fafc")
        pf.pack(fill="x", padx=8, pady=(4, 2))
        self.progress = ttk.Progressbar(pf, mode="determinate", style="Modern.Horizontal.TProgressbar")
        self.progress.pack(side="left", fill="x", expand=True)
        self.progress_label = tk.Label(pf, text="就绪", font=("Microsoft YaHei", 9),
                                       bg="#f8fafc", fg="#475569", width=12, anchor="e")
        self.progress_label.pack(side="left", padx=8)

        # 日志区卡片
        log_card = tk.Frame(self.root, bg="white", bd=0,
                            highlightbackground="#e2e8f0", highlightthickness=1)
        log_card.pack(fill="both", side="bottom", padx=8, pady=(4, 2))
        lh = tk.Frame(log_card, bg="#f1f5f9", height=28)
        lh.pack(fill="x")
        lh.pack_propagate(False)
        tk.Label(lh, text="  操作日志", font=("Microsoft YaHei", 10, "bold"),
                 bg="#f1f5f9", fg="#334155").pack(side="left", padx=4)
        self.log = scrolledtext.ScrolledText(log_card, height=8,
                                             font=("Consolas", 9), state="disabled",
                                             bg="#0f172a", fg="#e2e8f0",
                                             insertbackground="#e2e8f0",
                                             relief="flat", bd=0)
        self.log.pack(fill="both", expand=True, padx=2, pady=2)

    def _build_statusbar(self):
        sb = tk.Frame(self.root, bg="#1e293b", height=24)
        sb.pack(fill="x", side="bottom")
        sb.pack_propagate(False)
        self.status_var = tk.StringVar(value="就绪")
        tk.Label(sb, textvariable=self.status_var, anchor="w",
                 font=("Microsoft YaHei", 9), bg="#1e293b", fg="#e2e8f0").pack(side="left", padx=10)
        self.sched_var = tk.StringVar(value="")
        tk.Label(sb, textvariable=self.sched_var, anchor="e",
                 font=("Microsoft YaHei", 9), bg="#1e293b", fg="#94a3b8").pack(side="right", padx=10)

    # ====================== 设备列表操作 ======================
    def _load_devices(self):
        path = self.config.get("devices_file")
        self.devices = load_devices_json(path)
        self._log(f"已加载 {len(self.devices)} 台设备 ({path})")

    def _save_list(self):
        path = self.config.get("devices_file")
        save_devices_json(self.devices, path)
        self._log(f"设备清单已保存 ({path})")
        self.status_var.set(f"已保存 {len(self.devices)} 台设备")

    def _new_list(self):
        if self.devices and not messagebox.askyesno("确认", "清空当前设备清单?"):
            return
        self.devices = []
        self._refresh_table()
        self._log("已新建空清单")

    def _open_list(self):
        p = filedialog.askopenfilename(
            title="打开设备清单",
            filetypes=[("JSON 文件", "*.json")],
            initialdir=APP_DIR)
        if p:
            self.config.set("devices_file", p)
            self.config.save()
            self._load_devices()
            self._refresh_table()

    def _refresh_table(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        # 渲染时按状态+IP排序显示 (成功IP升序, 失败IP升序, 其余IP升序)
        display_devices = sort_devices_by_status_and_ip(self.devices)
        for i, d in enumerate(display_devices):
            if d.status == "成功":
                tag = "ok"
            elif d.status == "失败":
                tag = "fail"
            elif d.status in ("备份中", "排队中"):
                tag = "running"
            else:
                tag = "even" if i % 2 == 0 else ""
            # 状态用图标前缀增强可读性
            status_disp = {
                "成功": "● 成功",
                "失败": "● 失败",
                "备份中": "◐ 备份中",
                "排队中": "○ 排队中",
                "未备份": "○ 未备份",
            }.get(d.status, d.status)
            real_idx = self.devices.index(d)
            self.tree.insert("", "end", iid=str(real_idx),
                             values=(d.name or d.real_hostname or d.host,
                                     d.host, VENDOR_LABELS.get(d.vendor, d.vendor),
                                     d.protocol.upper(), d.port,
                                     d.group, status_disp, d.last_time, d.message),
                             tags=(tag,))
        # 更新统计
        total = len(self.devices)
        ok = sum(1 for d in self.devices if d.status == "成功")
        fail = sum(1 for d in self.devices if d.status == "失败")
        self.stat_var.set(f"共 {total} 台  ✓{ok}  ✗{fail}")

    def _add_device(self):
        dlg = DeviceDialog(self.root, config=self.config)
        if not dlg.result:
            return
        self._merge_devices(dlg.result)

    def _edit_device(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("提示", "请先选择一台设备")
            return
        idx = int(sel[0])
        dlg = DeviceDialog(self.root, self.devices[idx], "编辑设备", config=self.config)
        # 编辑模式 result 一定是单个 Device
        if dlg.result and not isinstance(dlg.result, list):
            # 保留运行时字段
            dlg.result.status = self.devices[idx].status
            dlg.result.last_time = self.devices[idx].last_time
            dlg.result.last_file = self.devices[idx].last_file
            dlg.result.real_hostname = self.devices[idx].real_hostname
            self.devices[idx] = dlg.result
            self._refresh_table()
            self._save_list_silent()

    def _merge_devices(self, new_devs):
        """合并设备(单个或列表)到当前列表, 按 IP 去重"""
        if isinstance(new_devs, Device):
            new_devs = [new_devs]
        if not new_devs:
            messagebox.showwarning("提示", "未导入任何设备")
            return
        existing_ips = {d.host for d in self.devices}
        added = 0
        for d in new_devs:
            if d.host not in existing_ips:
                self.devices.append(d)
                existing_ips.add(d.host)
                added += 1
        self._refresh_table()
        self._save_list_silent()
        skipped = len(new_devs) - added
        self._log(f"添加设备: 新增 {added} 台, 跳过重复 {skipped} 台, 当前列表共 {len(self.devices)} 台")
        # 无论是否新增都弹提示, 让用户知道操作结果
        if added:
            messagebox.showinfo("完成",
                f"已添加 {added} 台设备\n重复跳过 {skipped} 台\n当前列表共 {len(self.devices)} 台")
        else:
            messagebox.showinfo("提示",
                f"导入的 {len(new_devs)} 台设备 IP 均已存在, 全部跳过\n当前列表共 {len(self.devices)} 台")

    def _del_device(self):
        sel = self.tree.selection()
        if not sel:
            return
        if not messagebox.askyesno("确认", "删除选中的设备?"):
            return
        for s in sel:
            del self.devices[int(s)]
        self._refresh_table()
        self._save_list_silent()

    def _clear_all(self):
        """清空设备列表"""
        if not self.devices:
            return
        if not messagebox.askyesno("确认清空", f"确定清空全部 {len(self.devices)} 台设备?\n(已备份的配置文件不会被删除)"):
            return
        self.devices = []
        self._refresh_table()
        self._save_list_silent()
        self._log("已清空设备列表")

    def _batch_entry(self):
        """批量录入 - 打开添加设备对话框, 默认选中「批量录入」Tab"""
        dlg = DeviceDialog(self.root, config=self.config, initial_tab=1)
        if dlg.result:
            self._merge_devices(dlg.result)

    def _save_list_silent(self):
        save_devices_json(self.devices, self.config.get("devices_file"))

    # ====================== 批量导入 ======================
    def _import_file(self):
        """批量导入 - 打开添加设备对话框, 默认选中「从文件导入」Tab"""
        dlg = DeviceDialog(self.root, config=self.config, initial_tab=2)
        if dlg.result:
            self._merge_devices(dlg.result)

    def _unified_import(self):
        """统一密码导入 - 打开添加设备对话框, 默认选中「统一密码导入」Tab"""
        dlg = DeviceDialog(self.root, config=self.config, initial_tab=3)
        if dlg.result:
            self._merge_devices(dlg.result)

    def _export_template(self):
        p = filedialog.asksaveasfilename(
            title="保存导入模板",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv")],
            initialdir=APP_DIR,
            initialfile="devices_template.xlsx")
        if not p:
            return
        try:
            export_template(p)
            messagebox.showinfo("成功", f"模板已保存:\n{p}")
            self._log(f"导出模板: {p}")
        except Exception as e:
            messagebox.showerror("失败", str(e))

    # ====================== 备份 ======================
    def _backup_all(self):
        self._start_backup(self.devices)

    def _backup_selected(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("提示", "请先选择设备")
            return
        devs = [self.devices[int(s)] for s in sel]
        self._start_backup(devs)

    def _start_backup(self, devs: list):
        if self._is_backup_running:
            messagebox.showwarning("提示", "备份正在进行中")
            return
        if not devs:
            messagebox.showwarning("提示", "没有可备份的设备")
            return
        # 重置状态
        for d in devs:
            d.status = "排队中"
            d.message = ""
        self._refresh_table()
        self.progress["value"] = 0
        self.progress["maximum"] = len(devs)
        self.progress_label.config(text=f"0 / {len(devs)}")
        self._is_backup_running = True
        self.status_var.set("备份中...")

        os.makedirs(self.config.get("backup_dir"), exist_ok=True)
        self.engine = BackupEngine(
            backup_dir=self.config.get("backup_dir"),
            max_workers=self.config.get("max_workers", 10),
            on_progress=self._on_progress,
            on_log=self._on_log_thread,
        )
        self._backup_thread = threading.Thread(
            target=self._run_backup, args=(list(devs),), daemon=True)
        self._backup_thread.start()

    def _run_backup(self, devs):
        try:
            results = self.engine.backup_batch(devs)
        except Exception as e:
            self.root.after(0, lambda: self._log(f"备份线程异常: {e}"))
            results = devs
        self.root.after(0, lambda: self._on_backup_done(results))

    def _stop_backup(self):
        if self.engine and self._is_backup_running:
            self.engine.stop()
            self._log("正在停止备份...")

    def _on_progress(self, device: Device):
        # 来自子线程, 用 after 更新 UI
        self.root.after(0, lambda: self._update_device_row(device))

    def _update_device_row(self, device: Device):
        for i, d in enumerate(self.devices):
            if d is device or (d.host == device.host and d.port == device.port):
                # 用运行时数据覆盖
                d.status = device.status
                d.message = device.message
                d.last_time = device.last_time
                d.last_file = device.last_file
                d.real_hostname = device.real_hostname
                # 更新进度
                if device.status in ("成功", "失败"):
                    self.progress["value"] = self.progress["value"] + 1
                    self.progress_label.config(
                        text=f"{int(self.progress['value'])} / {int(self.progress['maximum'])}")
                break
        self._refresh_table()

    def _on_log_thread(self, msg):
        self.root.after(0, lambda: self._log(msg))

    def _on_backup_done(self, results):
        self._is_backup_running = False
        self.progress["value"] = self.progress["maximum"]
        ok = sum(1 for d in results if d.status == "成功")
        fail = len(results) - ok
        self.status_var.set(f"备份完成: 成功 {ok}, 失败 {fail}")
        self._log(f"备份结束: 成功 {ok}, 失败 {fail}")
        # 汇总日志: 成功列表(IP升序) + 失败列表(IP升序)
        sorted_results = sort_devices_by_status_and_ip(results)
        ok_list = [d for d in sorted_results if d.status == "成功"]
        fail_list = [d for d in sorted_results if d.status == "失败"]
        if ok_list:
            self._log("---- 成功设备 (按IP排序) ----")
            for d in ok_list:
                self._log(f"  ✓ {d.host}  {d.real_hostname or d.name or ''}")
        if fail_list:
            self._log("---- 失败设备 (按IP排序) ----")
            for d in fail_list:
                self._log(f"  ✗ {d.host}  {d.message}")
        self._save_list_silent()

        # 自动导出报告 (成功/失败分两个 Sheet)
        try:
            report_path = self._auto_export_report(results)
            if report_path:
                self._log(f"报告已自动导出: {report_path}")
        except Exception as e:
            self._log(f"自动导出报告失败: {e}")

        # 邮件通知
        if self.config.get("mail_enabled"):
            threading.Thread(target=self._send_mail, args=(results,), daemon=True).start()

    def _send_mail(self, results):
        try:
            html = build_report_html(f"交换机配置备份报告 - {datetime.now():%Y-%m-%d %H:%M}",
                                     results)
            ok, msg = send_report(
                smtp_host=self.config.get("smtp_host", ""),
                smtp_port=int(self.config.get("smtp_port", 465)),
                sender=self.config.get("mail_sender", ""),
                password=self.config.get("mail_password", ""),
                recipients=self.config.get_mail_recipients_list(),
                subject=f"交换机备份报告 - {datetime.now():%Y-%m-%d %H:%M}",
                html_body=html,
                use_ssl=self.config.get("smtp_ssl", True),
            )
            self.root.after(0, lambda: self._log(f"邮件通知: {msg}"))
        except Exception as e:
            self.root.after(0, lambda: self._log(f"邮件发送失败: {e}"))

    # ====================== 工具 ======================
    def _open_compare(self):
        CompareDialog(self.root, self.config.get("backup_dir"))

    def _open_search(self):
        SearchDialog(self.root, self.config.get("backup_dir"))

    def _open_backup_dir(self):
        d = self.config.get("backup_dir")
        os.makedirs(d, exist_ok=True)
        try:
            if sys.platform.startswith("win"):
                os.startfile(d)
            elif sys.platform == "darwin":
                import subprocess
                subprocess.Popen(["open", d])
            else:
                import subprocess
                subprocess.Popen(["xdg-open", d])
        except Exception as e:
            messagebox.showerror("错误", str(e))

    def _auto_export_report(self, results) -> str:
        """备份完成后自动导出报告到 reports/ 目录, 成功/失败分两个 Sheet

        返回文件路径, 失败返回空字符串
        """
        from openpyxl import Workbook

        report_dir = os.path.join(APP_DIR, "reports")
        os.makedirs(report_dir, exist_ok=True)
        filename = f"backup_report_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        filepath = os.path.join(report_dir, filename)

        sorted_devices = sort_devices_by_status_and_ip(results)
        ok_list = [d for d in sorted_devices if d.status == "成功"]
        fail_list = [d for d in sorted_devices if d.status == "失败"]
        other_list = [d for d in sorted_devices if d.status not in ("成功", "失败")]

        wb = Workbook()
        # Sheet1: 成功设备
        ws_ok = wb.active
        ws_ok.title = f"成功 ({len(ok_list)}台)"
        self._fill_report_sheet(ws_ok, ok_list)

        # Sheet2: 失败设备
        ws_fail = wb.create_sheet(f"失败 ({len(fail_list)}台)")
        self._fill_report_sheet(ws_fail, fail_list)

        # Sheet3: 其他状态 (如有)
        if other_list:
            ws_other = wb.create_sheet(f"其他 ({len(other_list)}台)")
            self._fill_report_sheet(ws_other, other_list)

        wb.save(filepath)
        return filepath

    def _fill_report_sheet(self, ws, devices):
        """填充一个报告 Sheet: 表头 + 数据行 + 列宽"""
        headers = ["设备名称", "IP", "厂商", "协议", "端口", "分组",
                   "状态", "备份时间", "实际主机名", "备份文件", "详情"]
        ws.append(headers)
        for d in devices:
            ws.append([
                d.name, d.host, VENDOR_LABELS.get(d.vendor, d.vendor),
                d.protocol, d.port, d.group,
                d.status, d.last_time, d.real_hostname, d.last_file, d.message
            ])
        for i in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + i)].width = 16
        ws.column_dimensions["K"].width = 40

    def _export_report(self):
        """手动导出报告 (弹窗选择路径, 成功/失败分两个 Sheet)"""
        p = filedialog.asksaveasfilename(
            title="导出备份报告",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialdir=APP_DIR,
            initialfile=f"backup_report_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
        if not p:
            return
        try:
            from openpyxl import Workbook

            sorted_devices = sort_devices_by_status_and_ip(self.devices)
            ok_list = [d for d in sorted_devices if d.status == "成功"]
            fail_list = [d for d in sorted_devices if d.status == "失败"]
            other_list = [d for d in sorted_devices if d.status not in ("成功", "失败")]

            wb = Workbook()
            ws_ok = wb.active
            ws_ok.title = f"成功 ({len(ok_list)}台)"
            self._fill_report_sheet(ws_ok, ok_list)

            ws_fail = wb.create_sheet(f"失败 ({len(fail_list)}台)")
            self._fill_report_sheet(ws_fail, fail_list)

            if other_list:
                ws_other = wb.create_sheet(f"其他 ({len(other_list)}台)")
                self._fill_report_sheet(ws_other, other_list)

            wb.save(p)
            messagebox.showinfo("成功", f"报告已导出:\n{p}")
            self._log(f"导出报告: {p} (成功{len(ok_list)}台 / 失败{len(fail_list)}台)")
        except Exception as e:
            messagebox.showerror("失败", str(e))

    # ====================== 设置 / 调度 ======================
    def _open_settings(self):
        if self._is_backup_running:
            messagebox.showwarning("提示", "备份进行中, 无法修改设置")
            return
        dlg = SettingsDialog(self.root, self.config)
        if dlg.result:
            self._log("设置已保存")
            self._apply_scheduler()

    def _apply_scheduler(self):
        enabled = self.config.get("schedule_enabled", False)
        if enabled and not self._is_backup_running:
            self._start_scheduler()
        else:
            self._stop_scheduler()

    def _start_scheduler(self):
        if self.scheduler and self.scheduler.is_running():
            self.scheduler.stop()
        self.scheduler = BackupScheduler(
            task_fn=self._scheduled_backup,
            mode=self.config.get("schedule_mode", "interval"),
            interval_hours=self.config.get("schedule_interval_hours", 24),
            daily_hour=self.config.get("schedule_daily_hour", 2),
            on_log=self._on_log_thread,
        )
        self.scheduler.start()
        self._update_sched_status()

    def _stop_scheduler(self):
        if self.scheduler and self.scheduler.is_running():
            self.scheduler.stop()
        self._update_sched_status()

    def _update_sched_status(self):
        if self.scheduler and self.scheduler.is_running():
            self.sched_var.set(f"定时: 已启用 (下次: {self.scheduler.next_run_str()})")
        else:
            self.sched_var.set("定时: 未启用")

    def _scheduled_backup(self):
        """定时器触发的备份(在调度线程内执行)"""
        if self._is_backup_running:
            self._on_log_thread("定时触发时已有备份在运行, 跳过")
            return
        if not self.devices:
            self._on_log_thread("定时触发: 无设备")
            return
        self.root.after(0, lambda: self._start_backup(list(self.devices)))

    # ====================== 日志 ======================
    def _log(self, msg: str):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log.config(state="normal")
        self.log.insert("end", f"[{ts}] {msg}\n")
        self.log.see("end")
        self.log.config(state="disabled")

    # ====================== 其他 ======================
    def _about(self):
        messagebox.showinfo(
            "关于",
            APP_TITLE + "\n\n"
            "支持华为 / 华三 / 锐捷 交换机配置备份\n"
            "SSH / Telnet 登录, 多线程并行, 按主机名分区保存\n"
            "批量导入 · 配置对比 · 全文搜索 · 定时备份 · 邮件通知\n\n"
            "保存位置: " + self.config.get("backup_dir"),
            parent=self.root)

    def _on_close(self):
        if self._is_backup_running:
            if not messagebox.askyesno("确认", "备份正在进行, 确定退出?"):
                return
            if self.engine:
                self.engine.stop()
        if self.scheduler and self.scheduler.is_running():
            self.scheduler.stop()
        self._save_list_silent()
        self.root.destroy()


def _setup_style(style: ttk.Style):
    """配置现代扁平主题样式"""
    try:
        style.theme_use("clam")
    except Exception:
        pass

    # 通用配色
    BG = "#ffffff"
    FG = "#1e293b"
    ACCENT = "#2563eb"
    BORDER = "#cbd5e1"
    SELECT = "#dbeafe"

    style.configure(".", background=BG, foreground=FG,
                    font=("Microsoft YaHei", 9))
    style.configure("TFrame", background=BG)
    style.configure("TLabel", background=BG, foreground=FG)
    style.configure("TLabelframe", background=BG, foreground=FG,
                    bordercolor=BORDER)
    style.configure("TLabelframe.Label", background=BG, foreground="#334155",
                    font=("Microsoft YaHei", 10, "bold"))

    # Treeview
    style.configure("Treeview",
                    background=BG, foreground=FG, fieldbackground=BG,
                    rowheight=28, font=("Microsoft YaHei", 9),
                    borderwidth=0)
    style.configure("Treeview.Heading",
                    background="#1e3a8a", foreground="white",
                    font=("Microsoft YaHei", 10, "bold"),
                    relief="flat", borderwidth=0)
    style.map("Treeview.Heading",
              background=[("active", "#1e40af")])
    style.map("Treeview",
              background=[("selected", SELECT)],
              foreground=[("selected", FG)])

    # 进度条
    style.configure("Modern.Horizontal.TProgressbar",
                    background=ACCENT, troughcolor="#e2e8f0",
                    borderwidth=0, thickness=14)
    style.map("Modern.Horizontal.TProgressbar",
              background=[("active", "#1d4ed8")])

    # Combobox
    style.configure("TCombobox",
                    fieldbackground=BG, background=BG,
                    foreground=FG, bordercolor=BORDER,
                    arrowcolor=ACCENT, padding=4)
    style.map("TCombobox",
              fieldbackground=[("readonly", BG)],
              bordercolor=[("focus", ACCENT)])

    # Entry
    style.configure("TEntry",
                    fieldbackground=BG, foreground=FG,
                    bordercolor=BORDER, padding=4)
    style.map("TEntry",
              bordercolor=[("focus", ACCENT)])

    # Scrollbar
    style.configure("TScrollbar",
                    background="#e2e8f0", troughcolor="#f1f5f9",
                    borderwidth=0, arrowcolor="#475569")
    style.map("TScrollbar",
              background=[("active", "#cbd5e1")])

    # Notebook
    style.configure("TNotebook", background="#f8fafc", borderwidth=0)
    style.configure("TNotebook.Tab",
                    background="#e2e8f0", foreground="#475569",
                    padding=(14, 6), font=("Microsoft YaHei", 9))
    style.map("TNotebook.Tab",
              background=[("selected", BG)],
              foreground=[("selected", ACCENT)])


def main():
    root = tk.Tk()
    # 高 DPI 自适应(Windows)
    try:
        root.call("tk", "scaling", 1.3)
    except Exception:
        pass
    root.configure(bg="#f8fafc")
    style = ttk.Style()
    _setup_style(style)
    MainWindow(root)
    root.mainloop()


if __name__ == "__main__":
    main()
