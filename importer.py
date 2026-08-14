# -*- coding: utf-8 -*-
"""
importer.py
批量导入设备列表 - 支持 Excel(.xlsx/.xls) / CSV / TXT
表头兼容中英文, 列顺序不限; 表头识别失败时按 IP 正则回退识别
"""
import csv
import os
import re

from device import Device

# 表头别名映射 (中文 / 英文 / 缩写 -> 字段)
# 已尽量覆盖主流写法; 仍匹配不到时启用 IP 正则回退
HEADER_ALIASES = {
    "name": ["name", "设备名称", "名称", "设备名", "备注", "交换机", "交换机名",
             "交换机名称", "设备", "主机名", "hostname", "devicename", "device_name",
             "switch", "switchname", "节点名"],
    "host": ["host", "ip", "ip地址", "管理ip", "地址", "主机", "交换机ip", "设备ip",
             "管理地址", "交换机地址", "ip_address", "ipaddress", "switch_ip",
             "设备地址", "交换机管理ip", "管理ip地址", "ip_addr", "ipaddress(ip)",
             "交换机ip地址", "设备ip地址", "管理ip(ip)", "switch_ip_address",
             "device_ip", "mgmt_ip", "管理ip地址(ip)"],
    "port": ["port", "端口", "port号", "端口号", "ssh端口", "telnet端口"],
    "vendor": ["vendor", "厂商", "品牌", "设备类型", "设备厂商", "设备品牌",
               "manufacture", "manufacturer", "type"],
    "protocol": ["protocol", "协议", "连接方式", "连接协议", "access"],
    "username": ["username", "user", "用户名", "账号", "账户", "登录用户",
                 "loginuser", "login", "user_name", "uid"],
    "password": ["password", "pwd", "密码", "口令", "登录密码", "pass", "passwd"],
    "enable_password": ["enable_password", "enable", "特权密码", "enable密码",
                        "secret", "en密码", "特权口令", "enablepwd"],
    "group": ["group", "分组", "组", "区域", "机房", "位置", "区域分组"],
    "timeout": ["timeout", "超时", "超时时间", "超时(秒)"],
}

DEFAULT_PORTS = {"ssh": 22, "telnet": 23}

# IPv4 正则 - 用于表头匹配失败时回退识别 IP 列
IP_REGEX = re.compile(
    r"^(?:\d{1,3}\.){3}\d{1,3}$"
)


def _normalize(s) -> str:
    """归一化表头: 去空格(含全角)/转小写/去常见标点"""
    if s is None:
        return ""
    s = str(s).strip().lower()
    # 去半角空格和全角空格
    s = s.replace(" ", "").replace("\u3000", "")
    # 去常见分隔标点
    s = s.replace("-", "").replace("_", "").replace("/", "").replace("(", "").replace(")", "")
    return s


def _build_header_map(headers: list) -> dict:
    """返回 {field: column_index}"""
    mapping = {}
    norm_headers = [_normalize(h) for h in headers]
    # 归一化后的别名集合
    norm_aliases = {f: [_normalize(a) for a in aliases]
                    for f, aliases in HEADER_ALIASES.items()}
    for field, aliases in norm_aliases.items():
        for i, nh in enumerate(norm_headers):
            if nh in aliases or nh == field:
                mapping[field] = i
                break
    return mapping


def _fallback_find_host_col(rows: list) -> int:
    """表头匹配不到 host 列时, 扫描前若干行数据, 找出哪一列的内容像 IP 地址
    返回列索引, 找不到返回 -1
    """
    sample = rows[:min(20, len(rows))]
    best_col, best_score = -1, 0
    if not sample:
        return -1
    col_count = max(len(r) for r in sample)
    for c in range(col_count):
        hits = 0
        for r in sample:
            if c < len(r):
                val = str(r[c]).strip()
                if IP_REGEX.match(val):
                    hits += 1
        if hits > best_score:
            best_score, best_col = hits, c
    return best_col if best_score > 0 else -1


def _row_to_device(row: list, mapping: dict, host_col_override: int = -1) -> Device:
    """把一行数据转换为 Device
    host_col_override: 当表头匹配不到 host 列时, 由 IP 正则回退识别出的列索引
    """
    def get(field, default=""):
        if field in mapping:
            val = row[mapping[field]] if mapping[field] < len(row) else default
            return str(val).strip() if val is not None else default
        return default

    # host 优先用表头映射, 否则用回退列
    if "host" in mapping and mapping["host"] < len(row):
        host = str(row[mapping["host"]]).strip() if row[mapping["host"]] is not None else ""
    elif host_col_override >= 0 and host_col_override < len(row):
        host = str(row[host_col_override]).strip() if row[host_col_override] is not None else ""
    else:
        host = ""

    vendor = get("vendor", "auto").lower()
    if vendor in ("华为", "hw"):
        vendor = "huawei"
    elif vendor in ("华三", "h3c", "hp"):
        vendor = "h3c"
    elif vendor in ("锐捷", "ruijie", "rj"):
        vendor = "ruijie"
    elif vendor in ("自动", "auto", "auto识别", "自动识别", ""):
        vendor = "auto"

    protocol = get("protocol", "ssh").lower()
    if protocol not in ("ssh", "telnet"):
        protocol = "ssh"

    port = get("port", "")
    try:
        port = int(port) if port else DEFAULT_PORTS[protocol]
    except ValueError:
        port = DEFAULT_PORTS[protocol]

    timeout = get("timeout", "")
    try:
        timeout = int(timeout) if timeout else 30
    except ValueError:
        timeout = 30

    return Device(
        name=get("name", ""),
        host=host,
        port=port,
        vendor=vendor,
        protocol=protocol,
        username=get("username", ""),
        password=get("password", ""),
        enable_password=get("enable_password", ""),
        group=get("group", "默认") or "默认",
        timeout=timeout,
    )


def import_devices(path: str) -> list:
    """从文件导入设备列表, 返回 Device 列表 (跳过无 host 的行)"""
    ext = os.path.splitext(path)[1].lower()
    devices = []

    if ext in (".xlsx", ".xls"):
        devices = _import_excel(path)
    elif ext == ".csv":
        devices = _import_csv(path)
    elif ext == ".txt":
        devices = _import_txt(path)
    else:
        raise ValueError(f"不支持的文件类型: {ext} (仅支持 xlsx/xls/csv/txt)")

    # 过滤掉无 IP 的行
    return [d for d in devices if d.host]


def _import_csv(path: str) -> list:
    devices = []
    reader = None
    # 尝试多种编码
    for enc in ("utf-8-sig", "utf-8", "gbk", "gb18030"):
        try:
            with open(path, "r", encoding=enc, newline="") as f:
                reader = list(csv.reader(f))
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("CSV 文件编码无法识别")

    if not reader:
        return devices

    mapping = _build_header_map(reader[0])
    # 表头没识别到 host 列时, 用 IP 正则回退定位
    host_col = -1
    if "host" not in mapping:
        host_col = _fallback_find_host_col(reader[1:])
    for row in reader[1:]:
        if not row or all(not str(c).strip() for c in row):
            continue
        devices.append(_row_to_device(row, mapping, host_col))
    return devices


def _import_txt(path: str) -> list:
    """TXT 格式: 每行一台设备, 字段用空格/逗号/制表符分隔, 第一行为表头
    或者简化格式: host username password [vendor]"""
    devices = []
    for enc in ("utf-8", "gbk", "gb18030"):
        try:
            with open(path, "r", encoding=enc) as f:
                lines = f.readlines()
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("TXT 文件编码无法识别")

    if not lines:
        return devices

    first = lines[0].strip()
    sep = "," if "," in first else ("\t" if "\t" in first else None)

    # 含表头(含字母 host/ip 等)
    if sep and any(k in first.lower() for k in ("host", "ip", "name", "交换机", "设备")):
        headers = [h.strip() for h in first.split(sep)]
        mapping = _build_header_map(headers)
        # 表头没识别到 host 列时, 用 IP 正则回退定位
        host_col = -1
        if "host" not in mapping:
            data_rows = []
            for line in lines[1:]:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                data_rows.append([c.strip() for c in line.split(sep)])
            host_col = _fallback_find_host_col(data_rows)
        for line in lines[1:]:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            row = [c.strip() for c in line.split(sep)]
            devices.append(_row_to_device(row, mapping, host_col))
    else:
        # 简化格式: host user password [vendor] [protocol]
        for line in lines:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            parts = line.replace(",", " ").split()
            if len(parts) < 3:
                continue
            host, user, pwd = parts[0], parts[1], parts[2]
            vendor = parts[3].lower() if len(parts) > 3 else "auto"
            protocol = parts[4].lower() if len(parts) > 4 else "ssh"
            if vendor in ("华为", "hw"):
                vendor = "huawei"
            elif vendor in ("华三", "h3c"):
                vendor = "h3c"
            elif vendor in ("锐捷", "rj"):
                vendor = "ruijie"
            elif vendor in ("自动", "auto"):
                vendor = "auto"
            port = 23 if protocol == "telnet" else 22
            devices.append(Device(host=host, username=user, password=pwd,
                                  vendor=vendor, protocol=protocol, port=port))
    return devices


def _import_excel(path: str) -> list:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    mapping = _build_header_map(list(rows[0]))
    # 表头没识别到 host 列时, 用 IP 正则回退定位
    host_col = -1
    if "host" not in mapping:
        host_col = _fallback_find_host_col([list(r) for r in rows[1:]])
    devices = []
    for row in rows[1:]:
        row = ["" if c is None else c for c in row]
        if all(str(c).strip() == "" for c in row):
            continue
        devices.append(_row_to_device(row, mapping, host_col))
    return devices


def export_template(path: str):
    """生成导入模板 (xlsx 或 csv)
    厂商列可填: auto(自动识别) / huawei / h3c / ruijie
    """
    headers = ["设备名称", "IP地址", "端口", "厂商", "协议", "用户名", "密码",
               "特权密码", "分组", "超时"]
    sample = [
        ["核心-自动识别", "10.0.0.1", 22, "auto", "ssh", "admin", "Admin@123", "", "核心", 30],
        ["接入-华为", "10.0.0.2", 22, "huawei", "ssh", "admin", "Admin@123", "", "接入", 30],
        ["接入-华三", "10.0.0.3", 22, "h3c", "ssh", "admin", "Admin@123", "", "接入", 30],
        ["接入-锐捷", "10.0.0.4", 23, "ruijie", "telnet", "admin", "Admin@123",
         "Admin@enable", "接入", 30],
    ]
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        for enc in ("utf-8-sig", "gbk"):
            try:
                with open(path, "w", encoding=enc, newline="") as f:
                    w = csv.writer(f)
                    w.writerow(headers)
                    w.writerows(sample)
                return
            except UnicodeEncodeError:
                continue
    else:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "设备列表"
        ws.append(headers)
        for s in sample:
            ws.append(s)
        # 列宽
        for i, h in enumerate(headers, 1):
            ws.column_dimensions[chr(64 + i)].width = 14
        wb.save(path)
